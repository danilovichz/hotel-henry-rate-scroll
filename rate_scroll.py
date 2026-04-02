#!/usr/bin/env python3
"""
Henry Rate Scroll — Automated Competitor Rate Monitoring
Hotel Henry (Holiday Inn Express & Suites San Diego Mission Valley)

Runs every 30 minutes via cron.
Writes to: daily Excel (.xlsx) + CSV backup
Alerts: Discord #henry channel when notable events occur

Data sources:
  - Expedia.com (public): lowest Expedia rate + room type breakdown (kings/queens count)
  - IHG.com: IHG direct rate for IHG-family hotels; Courtyard/Hampton get "X" (not IHG)

Room count logic (per Bobby, March 24 meeting):
  - If Expedia shows "X left" for a room type → use that number
  - If no scarcity indicator shown (ample rooms) → use hotel's arbitrary_rooms value
  - If sold out → 0
  - Total = kings + queens
  - Darshan to confirm arbitrary_rooms value per hotel (default: 20)

Strategy: HERD FOLLOWER — "We are a herd of hotels that can't be front runners." (Bobby)

Usage:
  python rate_scroll.py                    # Scrape today's rates (all hotels)
  python rate_scroll.py --date 2026-04-12  # Scrape a specific check-in date
  python rate_scroll.py --test             # Hampton Inn only, print result, no output writes
"""

import os
import re
import sys
import csv
import logging
import argparse
import requests
from datetime import date, timedelta, datetime
from zoneinfo import ZoneInfo
from pathlib import Path
from dotenv import load_dotenv

SD_TZ = ZoneInfo('America/Los_Angeles')

load_dotenv('/Users/rentamac/dani/aios/.env')

FIRECRAWL_API_KEY = os.getenv('FIRECRAWL_API_KEY')
FIRECRAWL_URL = 'https://api.firecrawl.dev/v1/scrape'
DISCORD_WEBHOOK = os.getenv('HENRY_DISCORD_WEBHOOK', '')

SCRIPT_DIR = Path(__file__).parent
LOG_DIR = SCRIPT_DIR / 'logs'
DATA_DIR = SCRIPT_DIR / 'data'
LOG_DIR.mkdir(exist_ok=True)
DATA_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
)
log = logging.getLogger(__name__)


# ─── Hotels ───────────────────────────────────────────────────────────────────
#
# expedia_id:    Expedia h-number (from URL: .h{id}.Hotel-Information)
# expedia_slug:  Full Expedia URL slug for URL construction
# ihg_code:      IHG.com 5-letter property code (None = not IHG family)
# ihg_family:    True if hotel is in IHG brand family (Holiday Inn, Crown Plaza, etc.)
# arbitrary_rooms: Fallback room count when Expedia shows no "X left" indicator
#                  Bobby: "if it's blank, count 20 rooms" — Darshan to confirm per hotel

OUR_HOTEL = {
    'name': 'Our Express HC',
    'expedia_id': 9303562,
    'expedia_slug': 'San-Diego-Hotels-Holiday-Inn-Express-Suites-San-Diego-Mission-Valley',
    'ihg_code': 'sanbp',
    'ihg_family': True,
    'is_ours': True,
    'arbitrary_rooms': 20,
}

COMP_HOTELS = [
    {
        'name': 'HIE SeaWorld',
        'expedia_id': 1676,
        'expedia_slug': 'San-Diego-Hotels-Holiday-Inn-Express-San-Diego-SeaWorld-Area',
        'ihg_code': 'sanex',   # Confirmed: IHG San Diego SeaWorld Area
        'ihg_family': True,
        'is_ours': False,
        'arbitrary_rooms': 20,
    },
    {
        'name': 'HIE Old Town',
        'expedia_id': 5735,
        'expedia_slug': 'San-Diego-Hotels-Holiday-Inn-Express-San-Diego-Airport-Old-Town',
        'ihg_code': 'sannf',   # Confirmed: IHG San Diego Airport - Old Town
        'ihg_family': True,
        'is_ours': False,
        'arbitrary_rooms': 20,
    },
    {
        'name': 'HIE Downtown',
        'expedia_id': 40681,
        'expedia_slug': 'San-Diego-Hotels-Holiday-Inn-Express-San-Diego-Downtown',
        'ihg_code': 'sanrx',   # Confirmed: IHG San Diego Downtown
        'ihg_family': True,
        'is_ours': False,
        'arbitrary_rooms': 20,
    },
    {
        'name': 'Courtyard',
        'expedia_id': 1987617,
        'expedia_slug': 'San-Diego-Hotels-Courtyard-By-Marriott-San-Diego-Mission-ValleyHotel-Circle',
        'ihg_code': None,
        'ihg_family': False,   # Marriott — IHG column = "X"
        'is_ours': False,
        'arbitrary_rooms': 20,
    },
    {
        'name': 'Hampton Inn',
        'expedia_id': 4415,
        'expedia_slug': 'San-Diego-Hotels-Hampton-Inn-San-Diego-Mission-Valley',
        'ihg_code': None,
        'ihg_family': False,   # Hilton — IHG column = "X"
        'is_ours': False,
        'arbitrary_rooms': 20,
    },
]

ALL_HOTELS = COMP_HOTELS + [OUR_HOTEL]  # Our hotel scraped last — lets Firecrawl warm up on comp pages first


# ─── URL Builders ─────────────────────────────────────────────────────────────

def expedia_url(hotel: dict, checkin: str, checkout: str) -> str:
    """Public Expedia.com hotel page with dates — shows room types, rates, and availability."""
    return (
        f"https://www.expedia.com/{hotel['expedia_slug']}.h{hotel['expedia_id']}.Hotel-Information"
        f"?chkin={checkin}&chkout={checkout}&adults=2"
    )


def build_ihg_url(hotel: dict, checkin: str, checkout: str):
    """IHG.com hotel page with dates — shows direct IHG rate and room availability."""
    code = hotel.get('ihg_code')
    if not code:
        return None
    ci = datetime.strptime(checkin, '%Y-%m-%d')
    co = datetime.strptime(checkout, '%Y-%m-%d')
    return (
        f"https://www.ihg.com/holidayinnexpress/hotels/us/en/san-diego/{code.lower()}/hoteldetail"
        f"?fromRedirect=true&qSrt=sBR&qSlH={code.upper()}"
        f"&qRms=1&qAdlt=2&qChld=0"
        f"&qCiD={ci.day}&qCiMy={ci.strftime('%m%Y')}"
        f"&qCoD={co.day}&qCoMy={co.strftime('%m%Y')}"
        f"&qpMbw=0"
    )


# ─── Firecrawl Schemas ────────────────────────────────────────────────────────

EXPEDIA_SCHEMA = {
    "type": "object",
    "properties": {
        "lowest_rate_usd": {
            "type": "number",
            "description": "Lowest available nightly rate in USD shown on this page. Null if sold out."
        },
        "is_sold_out": {
            "type": "boolean",
            "description": "True if the hotel has zero rooms available for this date."
        },
        "room_types": {
            "type": "array",
            "description": (
                "Every room type listed on the page. Capture ALL variants — "
                "e.g. '1 King Bed Non Smoking', '1 King Bed Accessible', '2 Queen Beds Non Smoking', etc."
            ),
            "items": {
                "type": "object",
                "properties": {
                    "name": {
                        "type": "string",
                        "description": "Exact room type name as shown on Expedia."
                    },
                    "rate_usd": {
                        "type": "number",
                        "description": "Nightly rate for this room type in USD."
                    },
                    "rooms_left": {
                        "type": "number",
                        "description": (
                            "How many rooms of this type are left. "
                            "If the page shows 'X left' or 'only X remaining' → use that number. "
                            "If no scarcity indicator shown (ample availability) → return null. "
                            "If this specific room type is sold out → return 0."
                        )
                    },
                    "is_sold_out": {
                        "type": "boolean",
                        "description": "True if this specific room type shows as sold out or unavailable."
                    }
                },
                "required": ["name", "is_sold_out"]
            }
        }
    },
    "required": ["is_sold_out"]
}

IHG_SCHEMA = {
    "type": "object",
    "properties": {
        "lowest_rate_usd": {
            "type": "number",
            "description": "Lowest available nightly rate in USD on IHG.com for this hotel. Null if sold out."
        },
        "is_sold_out": {
            "type": "boolean",
            "description": "True if no rooms are available."
        },
        "kings_available": {
            "type": "number",
            "description": (
                "Number of king rooms available. "
                "Use 'X left' indicator if shown. Return null if no scarcity shown. Return 0 if sold out."
            )
        },
        "queens_available": {
            "type": "number",
            "description": (
                "Number of queen/double rooms available. "
                "Use 'X left' indicator if shown. Return null if no scarcity shown. Return 0 if sold out."
            )
        }
    },
    "required": ["is_sold_out"]
}


# ─── Scraping ─────────────────────────────────────────────────────────────────

def _firecrawl_extract(url: str, schema: dict, prompt: str) -> dict:
    """Single Firecrawl LLM-extract call. Returns extracted dict or raises on failure."""
    payload = {
        "url": url,
        "formats": ["extract", "markdown"],  # markdown lets us detect bot challenges
        "extract": {"schema": schema, "prompt": prompt},
        "mobile": True,  # mobile UA is less likely to trigger bot detection
    }
    headers = {
        "Authorization": f"Bearer {FIRECRAWL_API_KEY}",
        "Content-Type": "application/json"
    }
    resp = requests.post(FIRECRAWL_URL, json=payload, headers=headers, timeout=60)
    resp.raise_for_status()
    data = resp.json()
    # Detect bot challenge — Expedia shows "Bot or Not?" CAPTCHA when blocking scrapers
    md = data.get('data', {}).get('markdown', '') or ''
    if 'Bot or Not' in md or 'You have been blocked' in md or 'human side' in md:
        raise ValueError('bot_challenge: Expedia bot detection triggered — result would be unreliable')
    if data.get('success') and data.get('data', {}).get('extract'):
        return data['data']['extract']
    raise ValueError(f"No extract returned: {data.get('error', 'unknown')}")


def _resolve_room_count(extracted_val, arbitrary: int, hotel_sold_out: bool = False) -> int:
    """
    Convert Expedia/IHG extracted room count to a usable integer.

    Bobby's rule (March 24 meeting):
      - If Expedia shows "X left" → use that number
      - If no scarcity message shown (ample rooms) → use arbitrary_rooms value
      - If sold out → 0

    LLM extraction problem: models sometimes return 0 instead of null for "not found".
    Safeguard: if the hotel has a rate (not fully sold out) and the count is 0 or null,
    treat it as "no indicator shown" and fall back to the arbitrary value.
    Only trust 0 when the whole hotel is confirmed sold out.
    """
    if hotel_sold_out:
        return 0
    # Positive value = explicit "X left" indicator was shown — trust it
    if extracted_val is not None and int(extracted_val) > 0:
        return int(extracted_val)
    # 0 or None = either "no scarcity shown" or LLM defaulted to 0 — use arbitrary
    return arbitrary


def scrape_hotel(hotel: dict, checkin: str, checkout: str) -> dict:
    """
    Scrape one hotel for a check-in date.

    Steps:
      1. Scrape Expedia.com public page → lowest Expedia rate + room counts (kings/queens)
      2. Scrape IHG.com page → IHG direct rate (IHG-family hotels only)
         Non-IHG hotels (Courtyard, Hampton) get ihg_rate_usd = 'X'

    Room count logic:
      - If Expedia shows "X left" → use that number
      - If no count shown (blank/ample) → use hotel's arbitrary_rooms value
      - If sold out → 0
      - Total = kings + queens
    """
    is_ihg_family = hotel.get('ihg_family', False)
    arbitrary = hotel.get('arbitrary_rooms', 20)

    base = {
        'hotel_name': hotel['name'],
        'checkin': checkin,
        'is_ours': hotel.get('is_ours', False),
        'ihg_family': is_ihg_family,
    }

    # ── 1. Expedia scrape ──────────────────────────────────────────────────────
    exp_url = expedia_url(hotel, checkin, checkout)
    exp_data = {}
    exp_error = None

    try:
        exp_data = _firecrawl_extract(
            exp_url,
            EXPEDIA_SCHEMA,
            (
                f"This is an Expedia.com hotel listing page for check-in {checkin}, check-out {checkout}. "
                f"Extract the lowest available nightly rate in USD. "
                f"List EVERY room type shown on the page with its exact name, nightly rate, and availability. "
                f"For rooms_left: use the exact number if the page shows 'X left' or 'only X remaining'. "
                f"If no scarcity indicator is shown for a room type, return null for rooms_left. "
                f"If a room type is sold out, set is_sold_out=true and rooms_left=0. "
                f"Set hotel-level is_sold_out=true only if ALL room types are unavailable."
            )
        )
        log.info(f"  Expedia OK — {hotel['name']}: ${exp_data.get('lowest_rate_usd')} "
                 f"({len(exp_data.get('room_types', []))} room types)")
    except Exception as e:
        exp_error = str(e)
        if 'bot_challenge' in str(e):
            log.warning(f"  Expedia BLOCKED — {hotel['name']}: Expedia bot check (will show as ERR, not SOLD OUT)")
        else:
            log.error(f"  Expedia FAIL — {hotel['name']}: {e}")

    # ── 2. IHG scrape ─────────────────────────────────────────────────────────
    ihg_rate = 'X'  # Default for non-IHG hotels
    ihg_error = None

    if is_ihg_family:
        ihg_rate = None  # Will be populated if we have a code
        ihg_url_str = build_ihg_url(hotel, checkin, checkout)

        if ihg_url_str:
            try:
                ihg_data = _firecrawl_extract(
                    ihg_url_str,
                    IHG_SCHEMA,
                    (
                        f"This is an IHG.com hotel booking page for check-in {checkin}, check-out {checkout}. "
                        f"Extract the lowest available nightly rate in USD (the IHG direct or member rate). "
                        f"Also count king and queen rooms available — use 'X left' if shown, null if no count shown, 0 if sold out. "
                        f"Set is_sold_out=true if no rooms are available."
                    )
                )
                if ihg_data.get('is_sold_out'):
                    ihg_rate = 'SOLD'
                else:
                    ihg_rate = ihg_data.get('lowest_rate_usd')
                log.info(f"  IHG OK — {hotel['name']}: {('SOLD OUT' if ihg_rate == 'SOLD' else f'${ihg_rate}')}")
            except Exception as e:
                ihg_error = str(e)
                log.error(f"  IHG FAIL — {hotel['name']}: {e}")
        else:
            log.info(f"  IHG SKIP — {hotel['name']}: no IHG code set (add to hotel config)")

    # ── 3. Room counts — sum from room_types list ─────────────────────────────
    # We sum per bed type from the full room list rather than trusting a single
    # LLM-aggregated number.  Expedia room names always contain "King" or "Queen".
    #
    # Examples of real Expedia names:
    #   "1 King Bed Non Smoking"              → king
    #   "1 King Bed Accessible Roll In Shower" → king
    #   "2 Queen Beds Non Smoking"             → queen
    #   "2 Queen Beds Accessible Roll In Shower Non Smoking" → queen
    #
    # rooms_left per type:
    #   explicit number → use it
    #   null → no scarcity indicator shown → use arbitrary_rooms
    #   0 / is_sold_out=True → sold out

    is_sold_out = exp_data.get('is_sold_out', False)
    room_types  = exp_data.get('room_types', [])

    kings_total  = 0
    queens_total = 0

    for rt in room_types:
        name_lower = (rt.get('name') or '').lower()
        rt_sold = rt.get('is_sold_out', False)
        rt_left = rt.get('rooms_left')

        count = _resolve_room_count(rt_left, arbitrary, rt_sold)

        # Use word-boundary matching to avoid "king" matching inside "smoking"
        is_king  = bool(re.search(r'\bking\b', name_lower))
        is_queen = bool(re.search(r'\b(queen|double)\b', name_lower))

        if is_king:
            kings_total += count
        elif is_queen:
            queens_total += count

    # If hotel is fully sold out override everything to 0
    if is_sold_out:
        kings_total  = 0
        queens_total = 0

    # If no room types were extracted at all (scrape returned empty list),
    # fall back to arbitrary for both rather than reporting 0
    if not room_types and not is_sold_out:
        kings_total  = arbitrary
        queens_total = arbitrary

    kings       = kings_total
    queens      = queens_total
    total_rooms = kings + queens

    return {
        **base,
        'lowest_rate_usd': exp_data.get('lowest_rate_usd'),
        'ihg_rate_usd': ihg_rate,
        'is_sold_out': is_sold_out,
        'kings_available': kings,
        'queens_available': queens,
        'total_rooms_available': total_rooms,
        'room_types': exp_data.get('room_types', []),
        'expedia_url': exp_url,
        'error': exp_error or ihg_error,
    }


# ─── Rate Scroll Run ──────────────────────────────────────────────────────────

def run_rate_scroll(checkin_date: date, hotels: list = None) -> list:
    """Scrape all hotels for a given check-in date. Returns list of result dicts."""
    if hotels is None:
        hotels = ALL_HOTELS

    checkin = checkin_date.strftime('%Y-%m-%d')
    checkout = (checkin_date + timedelta(days=1)).strftime('%Y-%m-%d')

    log.info(f"Rate scroll — Check-in: {checkin} | Hotels: {len(hotels)}")

    results = []
    for hotel in hotels:
        log.info(f"Scraping {hotel['name']}...")
        result = scrape_hotel(hotel, checkin, checkout)
        results.append(result)

    return results


# ─── Output: CSV ──────────────────────────────────────────────────────────────

CSV_COLUMNS = [
    'run_timestamp', 'checkin',
    'hotel_name', 'is_ours', 'ihg_family',
    'lowest_rate_usd', 'ihg_rate_usd',
    'is_sold_out',
    'kings_available', 'queens_available', 'total_rooms_available',
    'error', 'expedia_url'
]


def write_csv(results: list, run_time: datetime):
    """Append results to daily CSV. One file per day, one row per hotel per scrape run."""
    csv_path = DATA_DIR / f"rate_scroll_{run_time.strftime('%Y-%m-%d')}.csv"
    is_new = not csv_path.exists()

    with open(csv_path, 'a', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, extrasaction='ignore')
        if is_new:
            writer.writeheader()
        for r in results:
            row = {**r, 'run_timestamp': run_time.strftime('%Y-%m-%d %H:%M:%S')}
            writer.writerow(row)

    log.info(f"CSV: {len(results)} rows → {csv_path.name}")


# ─── Output: Excel (.xlsx) ────────────────────────────────────────────────────

# Row layout matching Darshan's "Rate Shop" template exactly
HOTEL_ROW_MAP = {
    'Our Express HC': 4,
    'HIE SeaWorld':   5,
    'HIE Old Town':   6,
    'HIE Downtown':   7,
    'Courtyard':      8,
    'Hampton Inn':    9,
}

ROW_LABELS = {
    1:  '',                               # Date — written as column header
    2:  '',                               # Day/time — written as column header
    3:  'Rooms Left to sell',
    4:  'Our Express HC',
    5:  'Holiday Inn Express Sea World',
    6:  'Holiday Inn Express Old Town',
    7:  'Holiday Inn Express Downtown',
    8:  'Courtyard Marriott Mission Valley',
    9:  'Hampton Inn Mission Valley',
    10: 'Out of order',
    11: 'Kings Open',
    12: 'Arrivals',
    13: 'Day Use',
    14: 'Early Departures',
    15: 'Hurdle Points',
    16: 'Declines',
    17: 'Front Desk',
}


def write_xlsx(results: list, run_time: datetime, checkin_date: date):
    """
    Write/update daily Excel file matching Darshan's Rate Shop format.
    One file per day. Each scrape run adds TWO columns: EXPEDIA + IHG.
    """
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("openpyxl not installed — skipping xlsx. Run: pip install openpyxl")
        return

    xlsx_path = DATA_DIR / f"Rate_Shop_{checkin_date.strftime('%Y-%m-%d')}.xlsx"
    day_name  = checkin_date.strftime('%A')
    time_str  = run_time.strftime('%I:%M%p').lstrip('0')
    stamp_str = f"{run_time.strftime('%-m/%-d')} {time_str}"  # e.g. "3/25 8:00AM"

    # Load or create workbook
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb.active
        exp_col = ws.max_column + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Rate Shop'
        exp_col = 2  # Column B = first EXPEDIA column

        # Column A: row labels
        ws.column_dimensions['A'].width = 38
        ws.cell(row=1, column=1, value=checkin_date.strftime('%m/%d/%Y')).font = Font(bold=True, name='Arial', size=10)
        ws.cell(row=2, column=1, value=f'DAY: {day_name}').font = Font(bold=True, name='Arial', size=10)
        for row_num, label in ROW_LABELS.items():
            if row_num >= 3 and label:
                cell = ws.cell(row=row_num, column=1, value=label)
                cell.font = Font(bold=True, name='Arial', size=10)
                cell.alignment = Alignment(vertical='center')

    ihg_col = exp_col + 1

    # ── Styles ────────────────────────────────────────────────────────────────
    header_fill = PatternFill('solid', fgColor='4472C4')
    header_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    data_font   = Font(name='Arial', size=10)
    sold_fill   = PatternFill('solid', fgColor='FF6B6B')
    sold_font   = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    avail_fill  = PatternFill('solid', fgColor='C6EFCE')
    limit_fill  = PatternFill('solid', fgColor='FFEB9C')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    ws.column_dimensions[get_column_letter(exp_col)].width = 14
    ws.column_dimensions[get_column_letter(ihg_col)].width = 14

    # ── Row 1-2: Headers ──────────────────────────────────────────────────────
    for col, label in ((exp_col, 'EXPEDIA'), (ihg_col, 'IHG')):
        c = ws.cell(row=1, column=col, value=label)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    for col in (exp_col, ihg_col):
        c = ws.cell(row=2, column=col, value=stamp_str)
        c.font = header_font; c.fill = header_fill
        c.alignment = Alignment(horizontal='center')

    # ── Rate data (rows 4-9) ──────────────────────────────────────────────────
    our_rooms_left = None

    for r in results:
        row_num = HOTEL_ROW_MAP.get(r.get('hotel_name', ''))
        if not row_num:
            continue

        is_sold = r.get('is_sold_out', False)
        exp_rate = r.get('lowest_rate_usd')
        ihg_rate = r.get('ihg_rate_usd')

        # EXPEDIA column
        ec = ws.cell(row=row_num, column=exp_col)
        ec.border = thin_border
        ec.alignment = Alignment(horizontal='center')
        if is_sold:
            ec.value = 'SOLD'; ec.fill = sold_fill; ec.font = sold_font
        elif exp_rate:
            ec.value = exp_rate; ec.font = data_font; ec.number_format = '$#,##0'
            ec.fill = avail_fill
        else:
            ec.value = 'N/A'; ec.font = data_font

        # IHG column
        ic = ws.cell(row=row_num, column=ihg_col)
        ic.border = thin_border
        ic.alignment = Alignment(horizontal='center')
        if not r.get('ihg_family'):
            ic.value = 'X'; ic.font = data_font
        elif ihg_rate == 'SOLD':
            ic.value = 'SOLD'; ic.fill = sold_fill; ic.font = sold_font
        elif isinstance(ihg_rate, (int, float)):
            ic.value = ihg_rate; ic.font = data_font; ic.number_format = '$#,##0'
        else:
            ic.value = '—'; ic.font = data_font  # No IHG code set yet

        if r.get('is_ours'):
            _raw = r.get('total_rooms_available')
            try:
                our_rooms_left = int(_raw) if _raw is not None else None
            except (ValueError, TypeError):
                our_rooms_left = None

    # ── Row 3: Rooms left to sell (OUR hotel) ─────────────────────────────────
    for col in (exp_col, ihg_col):
        rc = ws.cell(row=3, column=col)
        rc.border = thin_border
        rc.alignment = Alignment(horizontal='center')
        rc.font = Font(bold=True, name='Arial', size=11)
        if our_rooms_left is not None:
            rc.value = our_rooms_left
            if our_rooms_left <= 5:
                rc.fill = sold_fill
                rc.font = Font(bold=True, color='FFFFFF', name='Arial', size=11)
            elif our_rooms_left <= 15:
                rc.fill = limit_fill
            else:
                rc.fill = avail_fill

    # ── Manual rows 10-17: leave empty borders for Darshan ───────────────────
    for row_num in range(10, 18):
        for col in (exp_col, ihg_col):
            ws.cell(row=row_num, column=col).border = thin_border
            ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='center')
            ws.cell(row=row_num, column=col).font = data_font

    # ── Room Availability block (rows 19+) ────────────────────────────────────
    # One row per hotel showing Kings | Queens | Total at each time check.
    # Column layout per time check: EXPEDIA col = "K:X  Q:Y", IHG col = Total
    # Stays in sync with the rate block above (same column pairs grow rightward).
    #
    # Row layout:
    #   Row 19: "ROOM AVAILABILITY" section header (written once, col A only)
    #   Row 20: column sub-headers ("Kings / Queens" + "Total")  — written once
    #   Rows 21-26: one hotel per row, same order as rate block

    ROOM_ROW_MAP = {
        'Our Express HC': 21,
        'HIE SeaWorld':   22,
        'HIE Old Town':   23,
        'HIE Downtown':   24,
        'Courtyard':      25,
        'Hampton Inn':    26,
    }

    ROOM_ROW_LABELS = {
        19: 'ROOM AVAILABILITY',
        20: '',
        21: 'Our Express HC',
        22: 'Holiday Inn Express Sea World',
        23: 'Holiday Inn Express Old Town',
        24: 'Holiday Inn Express Downtown',
        25: 'Courtyard Marriott Mission Valley',
        26: 'Hampton Inn Mission Valley',
    }

    section_fill = PatternFill('solid', fgColor='1F3864')  # Dark navy
    section_font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    sub_fill     = PatternFill('solid', fgColor='2F5496')
    sub_font     = Font(bold=True, color='FFFFFF', name='Arial', size=9)
    rooms_font   = Font(name='Arial', size=10)
    sold_rooms_font = Font(bold=True, color='FF0000', name='Arial', size=10)

    # Write row labels on first run only (col A doesn't exist yet for new sheet)
    if exp_col == 2:
        for row_num, label in ROOM_ROW_LABELS.items():
            if label:
                c = ws.cell(row=row_num, column=1, value=label)
                if row_num == 19:
                    c.font = Font(bold=True, name='Arial', size=11)
                else:
                    c.font = Font(bold=True, name='Arial', size=10)
                c.alignment = Alignment(vertical='center')

    # Section header row 19: dark navy background across both columns
    for col in (exp_col, ihg_col):
        c19 = ws.cell(row=19, column=col)
        c19.fill = section_fill
        c19.border = thin_border

    # Sub-header row 20: "Kings / Queens" and "Total"
    kq_cell = ws.cell(row=20, column=exp_col, value='Kings / Queens')
    kq_cell.font = sub_font; kq_cell.fill = sub_fill
    kq_cell.alignment = Alignment(horizontal='center')
    kq_cell.border = thin_border

    tot_cell = ws.cell(row=20, column=ihg_col, value='Total')
    tot_cell.font = sub_font; tot_cell.fill = sub_fill
    tot_cell.alignment = Alignment(horizontal='center')
    tot_cell.border = thin_border

    # Hotel room data rows 21-26
    for r in results:
        row_num = ROOM_ROW_MAP.get(r.get('hotel_name', ''))
        if not row_num:
            continue

        def _int(v, default=0):
            try: return int(v)
            except (ValueError, TypeError): return default
        kings  = _int(r.get('kings_available', 0))
        queens = _int(r.get('queens_available', 0))
        total  = _int(r.get('total_rooms_available', 0))
        sold   = r.get('is_sold_out', False)

        # Kings / Queens cell  (e.g.  "K: 3 | Q: 2")
        kq_c = ws.cell(row=row_num, column=exp_col)
        kq_c.border = thin_border
        kq_c.alignment = Alignment(horizontal='center')
        if sold:
            kq_c.value = 'SOLD OUT'
            kq_c.font  = sold_rooms_font
        else:
            kq_c.value = f'K: {kings}   Q: {queens}'
            kq_c.font  = rooms_font
            if total <= 3:
                kq_c.fill = sold_fill
                kq_c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            elif total <= 8:
                kq_c.fill = limit_fill

        # Total cell
        tot_c = ws.cell(row=row_num, column=ihg_col)
        tot_c.border = thin_border
        tot_c.alignment = Alignment(horizontal='center')
        if sold:
            tot_c.value = 0
            tot_c.font  = sold_rooms_font
        else:
            tot_c.value = total
            tot_c.font  = Font(bold=True, name='Arial', size=10)
            if total <= 3:
                tot_c.fill = sold_fill
                tot_c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            elif total <= 8:
                tot_c.fill = limit_fill
            else:
                tot_c.fill = avail_fill

    wb.save(xlsx_path)
    log.info(f"XLSX: 2 columns added → {xlsx_path.name}")


def write_epc_sheet(epc_data: dict, target_date: date = None):
    """
    Add/overwrite an 'EPC Forward' sheet in today's daily Excel file.
    Shows 9-day forward rates from Expedia Partner Central Rev+.
    Never touches the existing 'Rate Shop' sheet.
    """
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("openpyxl not installed — skipping EPC sheet")
        return

    if target_date is None:
        target_date = datetime.now(SD_TZ).date()

    xlsx_path = DATA_DIR / f"Rate_Shop_{target_date.strftime('%Y-%m-%d')}.xlsx"

    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        # Remove old EPC sheet if it exists (we overwrite it each daily run)
        if 'EPC Forward' in wb.sheetnames:
            del wb['EPC Forward']
    else:
        wb = Workbook()
        wb.active.title = 'Rate Shop'  # placeholder so Rate Shop always exists

    ws = wb.create_sheet('EPC Forward')

    # ── Styles ────────────────────────────────────────────────────────────────
    header_fill  = PatternFill('solid', fgColor='1F3864')   # dark navy
    header_font  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    our_fill     = PatternFill('solid', fgColor='4472C4')   # blue — our hotel
    our_font     = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    avg_fill     = PatternFill('solid', fgColor='7F7F7F')   # gray — comp avg
    avg_font     = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    label_font   = Font(bold=True, name='Arial', size=10)
    data_font    = Font(name='Arial', size=10)
    min_font     = Font(italic=True, color='7F7F7F', name='Arial', size=9)
    sold_fill    = PatternFill('solid', fgColor='FF6B6B')
    sold_font    = Font(bold=True, color='FFFFFF', name='Arial', size=9)
    thin_border  = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    # ── Column A: labels ──────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 36

    # ── Row 1: sheet title ────────────────────────────────────────────────────
    ts = epc_data.get('timestamp', '')[:16].replace('T', ' ')
    title_cell = ws.cell(row=1, column=1, value=f'EPC FORWARD RATES  (pulled {ts} SD)')
    title_cell.font = Font(bold=True, name='Arial', size=11)

    # ── Get ordered dates from our_hotel or comp_set_avg ─────────────────────
    dates = sorted(
        set(epc_data.get('our_hotel', {}).keys())
        | set(epc_data.get('comp_set_avg', {}).keys())
    )
    if not dates:
        log.warning("EPC sheet: no dates found in data — skipping")
        wb.save(xlsx_path)
        return

    # ── Row 2: date headers ───────────────────────────────────────────────────
    for col_offset, d in enumerate(dates):
        col = col_offset + 2  # B onwards
        ws.column_dimensions[get_column_letter(col)].width = 11
        dt = date.fromisoformat(d)
        label = dt.strftime('%-m/%-d\n%a')  # e.g. "3/25\nWed"
        c = ws.cell(row=2, column=col, value=label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = thin_border

    # Row 2 label cell
    c = ws.cell(row=2, column=1, value='Hotel')
    c.font = header_font; c.fill = header_fill; c.border = thin_border

    # ── Helper: write a rate row ──────────────────────────────────────────────
    def write_rate_row(row_num, name, rates_dict, label_fill=None, label_font_=label_font):
        c = ws.cell(row=row_num, column=1, value=name)
        c.font = label_font_
        c.border = thin_border
        if label_fill:
            c.fill = label_fill

        for col_offset, d in enumerate(dates):
            col = col_offset + 2
            val = rates_dict.get(d)
            cell = ws.cell(row=row_num, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            if val is None:
                cell.value = '—'; cell.font = min_font
            elif val == 0:
                cell.value = 'SOLD'; cell.fill = sold_fill; cell.font = sold_font
            elif val == 'min_stay':
                cell.value = 'Min Stay'; cell.font = min_font
            elif isinstance(val, int):
                cell.value = val
                cell.number_format = '$#,##0'
                cell.font = label_font_ if label_fill else data_font
            else:
                cell.value = str(val); cell.font = data_font

    # ── Our hotel ─────────────────────────────────────────────────────────────
    write_rate_row(3, 'Our Express HC', epc_data.get('our_hotel', {}),
                   label_fill=our_fill, label_font_=our_font)

    # ── Comp set average ──────────────────────────────────────────────────────
    write_rate_row(4, 'Comp Set Avg (EPC)', epc_data.get('comp_set_avg', {}),
                   label_fill=avg_fill, label_font_=avg_font)

    # ── Blank separator ───────────────────────────────────────────────────────
    ws.cell(row=5, column=1, value='').border = thin_border

    # ── Individual competitors ────────────────────────────────────────────────
    comp_order = [
        'HIE SeaWorld', 'HIE Old Town', 'HIE Downtown',
        'Courtyard', 'Hampton Inn', 'Hilton Garden Inn',
        'DoubleTree', 'Fairfield Inn', 'Best Western',
        'Candlewood Suites', 'Legacy Resort',
    ]
    row_num = 6
    for comp_name in comp_order:
        comp_data = epc_data.get('competitors', {}).get(comp_name)
        if comp_data is not None:
            write_rate_row(row_num, comp_name, comp_data)
            row_num += 1

    # ── Events section ────────────────────────────────────────────────────────
    events = epc_data.get('events', [])
    if events:
        evt_row = row_num + 1  # one blank row after competitors

        # Section header
        evt_header = ws.cell(row=evt_row, column=1, value='UPCOMING EVENTS')
        evt_header.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        evt_header.fill = PatternFill('solid', fgColor='1F3864')
        evt_header.border = thin_border

        # Column headers
        evt_row += 1
        for col, heading in enumerate(['Category', 'Event', 'Dates', 'Attendees'], start=1):
            c = ws.cell(row=evt_row, column=col)
            c.value = heading
            c.font = Font(bold=True, color='FFFFFF', name='Arial', size=9)
            c.fill = PatternFill('solid', fgColor='2F5496')
            c.border = thin_border
            c.alignment = Alignment(horizontal='center')

        # Set column widths for event columns
        ws.column_dimensions['B'].width = max(ws.column_dimensions['B'].width, 40)
        ws.column_dimensions['C'].width = max(ws.column_dimensions['C'].width, 36)
        ws.column_dimensions['D'].width = max(ws.column_dimensions['D'].width, 14)

        # Event rows
        for evt in events:
            evt_row += 1
            ws.cell(row=evt_row, column=1, value=evt.get('category', '')).font = data_font
            ws.cell(row=evt_row, column=2, value=evt.get('name', '')).font = data_font
            ws.cell(row=evt_row, column=3, value=evt.get('dates', '')).font = data_font
            att_cell = ws.cell(row=evt_row, column=4, value=evt.get('attendees'))
            att_cell.font = data_font
            att_cell.number_format = '#,##0'
            att_cell.alignment = Alignment(horizontal='right')
            for col in range(1, 5):
                ws.cell(row=evt_row, column=col).border = thin_border

    wb.save(xlsx_path)
    log.info(f"EPC sheet written → {xlsx_path.name} (EPC Forward tab, {len(dates)} dates, {len(events)} events)")


# ─── Data Quality Thresholds ──────────────────────────────────────────────────

RATE_FLOOR       = 50    # Below this is almost certainly a scrape error
RATE_CEILING     = 700   # Above this is suspicious for SD Mission Valley hotels
RATE_DELTA_PCT   = 40    # % change vs previous run that triggers a flag
OUR_PARITY_GAP   = 25    # % gap between our Expedia vs IHG rate (parity issue)
FALLBACK_TRIGGER = 3     # How many hotels showing K:20 Q:20 before flagging the run


# ─── Data Quality Checks ──────────────────────────────────────────────────────

def check_data_quality(results: list, prev_results: dict) -> list:
    """
    Validate scraped data for internal consistency and suspicious values.
    Returns a list of flag strings. Empty = clean run.

    Checks:
      1. Has rate but marked sold out (contradictory)
      2. Not sold out, no rate, no error (failed extraction)
      3. Rate below floor (likely wrong)
      4. Rate above ceiling (unusually high — verify)
      5. Rate changed >40% vs previous run (jump too big for 30 min)
      6. Our hotel: Expedia vs IHG parity gap >25% (rate parity issue)
      7. 3+ hotels showing fallback room counts (K:20 Q:20) in same run
    """
    flags = []

    for r in results:
        name     = r.get('hotel_name', 'Unknown')
        exp_rate = r.get('lowest_rate_usd')
        ihg_rate = r.get('ihg_rate_usd')
        is_sold  = r.get('is_sold_out', False)
        is_ours  = r.get('is_ours', False)

        # 1. Contradictory: has rate but sold out flag set
        if is_sold and exp_rate:
            flags.append(
                f"⚠️ **{name}**: marked SOLD OUT but Expedia shows ${exp_rate:.0f} — likely page error"
            )

        # 2. Not sold out, no rate extracted, no error reported
        if not is_sold and exp_rate is None and not r.get('error'):
            flags.append(
                f"⚠️ **{name}**: no Expedia rate returned (not sold out) — page may have loaded wrong"
            )

        # 3. Rate below floor
        if exp_rate and exp_rate < RATE_FLOOR:
            flags.append(
                f"⚠️ **{name}**: Expedia ${exp_rate:.0f} is suspiciously low (floor: ${RATE_FLOOR}) — verify"
            )

        # 4. Rate above ceiling
        if exp_rate and exp_rate > RATE_CEILING:
            flags.append(
                f"⚠️ **{name}**: Expedia ${exp_rate:.0f} is unusually high — verify manually"
            )

        # 5. Rate delta vs previous run
        prev_rate_raw = prev_results.get(name, {}).get('lowest_rate_usd', '')
        if prev_rate_raw:
            try:
                prev_rate = float(prev_rate_raw)
                if exp_rate and prev_rate > 0:
                    delta_pct = abs(exp_rate - prev_rate) / prev_rate * 100
                    if delta_pct > RATE_DELTA_PCT:
                        direction = "↑" if exp_rate > prev_rate else "↓"
                        flags.append(
                            f"⚠️ **{name}**: Expedia rate {direction}{delta_pct:.0f}% in one run "
                            f"(${prev_rate:.0f} → ${exp_rate:.0f}) — verify"
                        )
            except (ValueError, ZeroDivisionError):
                pass

        # 6. Our hotel: rate parity gap between Expedia and IHG direct
        if is_ours and exp_rate and isinstance(ihg_rate, (int, float)):
            gap_pct = abs(exp_rate - ihg_rate) / max(exp_rate, ihg_rate) * 100
            if gap_pct > OUR_PARITY_GAP:
                lower = "Expedia" if exp_rate < ihg_rate else "IHG"
                flags.append(
                    f"⚠️ **Rate parity gap — our hotel**: Expedia ${exp_rate:.0f} vs IHG ${ihg_rate:.0f} "
                    f"({gap_pct:.0f}% apart, {lower} is lower) — intentional?"
                )

    # 7. Multiple hotels showing fallback room count (whole-run scrape issue)
    fallback_hotels = [
        r.get('hotel_name') for r in results
        if not r.get('is_sold_out')
        and r.get('kings_available') == 20
        and r.get('queens_available') == 20
    ]
    if len(fallback_hotels) >= FALLBACK_TRIGGER:
        flags.append(
            f"⚠️ **{len(fallback_hotels)} hotels showing fallback room counts** (K:20 Q:20): "
            f"{', '.join(fallback_hotels)} — Expedia scarcity data may not have loaded this run"
        )

    return flags


def send_discord_flags(flags: list, run_time: datetime):
    """Send data quality flags to Discord — separate from business alerts."""
    if not DISCORD_WEBHOOK:
        for f in flags:
            log.warning(f"DATA FLAG: {f}")
        return
    message = (
        f"**Henry — Data Flags** `{run_time.strftime('%I:%M %p')}`\n"
        f"*Potential scrape issues — verify before acting on these numbers.*\n\n"
        + "\n".join(flags)
    )
    try:
        requests.post(DISCORD_WEBHOOK, json={"content": message}, timeout=10).raise_for_status()
        log.info(f"Discord: {len(flags)} data flag(s) sent")
    except Exception as e:
        log.error(f"Discord data flags failed: {e}")


# ─── Alerts: Discord ──────────────────────────────────────────────────────────

def load_previous_results(today: str) -> dict:
    """Load the penultimate scrape run from today's CSV for change detection."""
    csv_path = DATA_DIR / f"rate_scroll_{today}.csv"
    if not csv_path.exists():
        return {}

    rows = []
    with open(csv_path) as f:
        rows = list(csv.DictReader(f))
    if not rows:
        return {}

    seen = []
    grouped = {}
    for r in rows:
        ts = r['run_timestamp']
        if ts not in grouped:
            grouped[ts] = []
            seen.append(ts)
        grouped[ts].append(r)

    if len(seen) < 2:
        return {}

    prev_batch = grouped[seen[-2]]
    return {r['hotel_name']: r for r in prev_batch}


def check_alerts(results: list, prev_results: dict) -> list:
    """
    Alert logic — HERD FOLLOWER strategy.
    Bobby: "We are a herd of hotels that can't be front runners."

    DO alert on:
      - Comp just went sold out (hold/raise opportunity)
      - 2+ comps dropped rates (market moving down, consider following)
      - 15+ rooms unsold after 6pm (late inventory risk)
      - Comp kings just sold out

    NEVER alert "comp is higher than us, raise rate" — WRONG for herd hotel
    NEVER recommend raising when we have 20+ rooms unsold
    """
    alerts = []
    our_result = next((r for r in results if r.get('is_ours')), None)
    our_rooms = our_result.get('total_rooms_available') if our_result else None
    sd_now = datetime.now(SD_TZ)
    is_evening = sd_now.hour >= 18

    # 1. Comp just went sold out
    for r in results:
        if r.get('is_ours'):
            continue
        name = r.get('hotel_name', '')
        prev = prev_results.get(name, {})
        was_sold = prev.get('is_sold_out', 'False') == 'True'
        if r.get('is_sold_out') and not was_sold:
            rooms_str = f" We have {our_rooms} rooms." if our_rooms else ""
            alerts.append(f"🔴 **{name}** just SOLD OUT.{rooms_str} Hold rate.")

    # 2. Market moving down (2+ comps dropped)
    drops = []
    for r in results:
        if r.get('is_ours'):
            continue
        name = r.get('hotel_name', '')
        rate = r.get('lowest_rate_usd')
        prev_rate_raw = prev_results.get(name, {}).get('lowest_rate_usd', '')
        prev_rate = float(prev_rate_raw) if prev_rate_raw else None
        if rate and prev_rate and rate < prev_rate:
            drops.append((name, prev_rate, rate))

    if len(drops) >= 2:
        detail = ", ".join([f"{n} ${r:.0f}" for n, _, r in drops])
        our_rate = our_result.get('lowest_rate_usd') if our_result else None
        our_str = f" We're at ${our_rate:.0f}." if our_rate else ""
        alerts.append(f"📉 Market moving down. {detail}.{our_str} Follow?")

    # 3. Late inventory risk (15+ rooms after 6pm)
    if is_evening and our_rooms and our_rooms >= 15:
        comp_rates = [r.get('lowest_rate_usd') for r in results
                      if not r.get('is_ours') and r.get('lowest_rate_usd')]
        if comp_rates:
            alerts.append(
                f"⚠️ {our_rooms} rooms unsold at {sd_now.strftime('%I:%M%p')}. "
                f"Market: ${min(comp_rates):.0f}–${max(comp_rates):.0f}. Consider adjusting."
            )

    # 4. Comp kings just sold out
    for r in results:
        if r.get('is_ours'):
            continue
        name = r.get('hotel_name', '')
        kings = r.get('kings_available')
        prev_kings_raw = prev_results.get(name, {}).get('kings_available', '')
        prev_kings = int(prev_kings_raw) if str(prev_kings_raw).isdigit() else None
        if kings == 0 and prev_kings and prev_kings > 0:
            rate = r.get('lowest_rate_usd')
            rate_str = f" at ${rate:.0f}" if rate else ""
            alerts.append(f"👑 **{name}** kings SOLD OUT{rate_str}.")

    return alerts


def send_discord_alert(alerts: list, run_time: datetime):
    """Send alert digest to Discord via webhook."""
    if not alerts or not DISCORD_WEBHOOK:
        if alerts:
            log.info("Alerts triggered but HENRY_DISCORD_WEBHOOK not set:\n" + "\n".join(alerts))
        return
    message = f"**Henry — Rate Alert** `{run_time.strftime('%I:%M %p')}`\n\n" + "\n".join(alerts)
    try:
        requests.post(DISCORD_WEBHOOK, json={"content": message}, timeout=10).raise_for_status()
        log.info(f"Discord: {len(alerts)} alert(s) sent")
    except Exception as e:
        log.error(f"Discord alert failed: {e}")


def send_discord_xlsx(run_time: datetime, checkin_date: date):
    """Upload the day's Rate Shop xlsx to Discord after each scrape run."""
    if not DISCORD_WEBHOOK:
        return
    xlsx_path = DATA_DIR / f"Rate_Shop_{checkin_date.strftime('%Y-%m-%d')}.xlsx"
    if not xlsx_path.exists():
        return
    time_str = run_time.strftime('%I:%M %p').lstrip('0')
    message = f"**Rate Shop** — {checkin_date.strftime('%m/%d')} | Updated {time_str}"
    try:
        with open(xlsx_path, 'rb') as f:
            requests.post(
                DISCORD_WEBHOOK,
                data={"content": message},
                files={"file": (xlsx_path.name, f,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
                timeout=15,
            ).raise_for_status()
        log.info(f"Discord: xlsx uploaded — {xlsx_path.name}")
    except Exception as e:
        log.error(f"Discord xlsx upload failed: {e}")


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Henry Rate Scroll — automated competitor rate monitoring")
    parser.add_argument('--date', help='Check-in date YYYY-MM-DD (default: today)')
    parser.add_argument('--test', action='store_true', help='Hampton Inn only, print results, no writes')
    parser.add_argument('--no-alerts', action='store_true', help='Skip Discord alerts this run')
    parser.add_argument('--force', action='store_true', help='Bypass operating hours check')
    parser.add_argument('--no-xlsx', action='store_true', help='Skip xlsx write and Discord upload')
    args = parser.parse_args()

    if not FIRECRAWL_API_KEY:
        log.error("FIRECRAWL_API_KEY not set in .env")
        sys.exit(1)

    # Operating hours: 8:00 AM – 2:30 AM San Diego time
    if not args.test and not args.date and not args.force:
        sd_now = datetime.now(SD_TZ)
        h, m = sd_now.hour, sd_now.minute
        in_hours = h >= 8 or h <= 1 or (h == 2 and m <= 30)
        if not in_hours:
            log.info(f"Outside operating hours ({sd_now.strftime('%I:%M %p')} SD). Skipping.")
            sys.exit(0)

    checkin_date = datetime.now(SD_TZ).date()
    if args.date:
        try:
            checkin_date = date.fromisoformat(args.date)
        except ValueError:
            log.error(f"Invalid date: {args.date}. Use YYYY-MM-DD.")
            sys.exit(1)

    hotels = [COMP_HOTELS[4]] if args.test else ALL_HOTELS  # Hampton only in test
    run_time = datetime.now(SD_TZ)

    results = run_rate_scroll(checkin_date, hotels)

    # ── Print summary ─────────────────────────────────────────────────────────
    print(f"\n{'=' * 80}")
    print(f"  Henry Rate Scroll — {run_time.strftime('%Y-%m-%d %H:%M')}  |  Check-in: {checkin_date}")
    print(f"{'=' * 80}")
    print(f"  {'Hotel':<26} {'Expedia':>9} {'IHG':>9}  {'Sold':>5}  {'Kings':>5} {'Queens':>6} {'Total':>5}")
    print(f"  {'-' * 72}")
    for r in results:
        exp_str = ('SOLD OUT' if r.get('is_sold_out')
                   else (f"${r['lowest_rate_usd']:.0f}" if r.get('lowest_rate_usd') else 'ERR' if r.get('error') else 'N/A'))
        ihg_val = r.get('ihg_rate_usd')
        ihg_str = ('X' if ihg_val == 'X'
                   else 'SOLD' if ihg_val == 'SOLD'
                   else f"${ihg_val:.0f}" if isinstance(ihg_val, (int, float)) else '—')
        sold  = '✓' if r.get('is_sold_out') else ''
        kings = str(r.get('kings_available', '--'))
        queens = str(r.get('queens_available', '--'))
        total = str(r.get('total_rooms_available', '--'))
        flag  = ' ← OURS' if r.get('is_ours') else ''
        print(f"  {r.get('hotel_name',''):<26} {exp_str:>9} {ihg_str:>9}  {sold:>5}  {kings:>5} {queens:>6} {total:>5}{flag}")
    print(f"{'=' * 80}\n")

    if args.test:
        log.info("Test mode — skipping all writes")
        return

    write_csv(results, run_time)
    if not args.no_xlsx:
        write_xlsx(results, run_time, checkin_date)
        send_discord_xlsx(run_time, checkin_date)

    if not args.no_alerts:
        prev = load_previous_results(run_time.strftime('%Y-%m-%d'))

        # Business alerts (herd follower logic)
        alerts = check_alerts(results, prev)
        if alerts:
            for a in alerts:
                log.info(f"ALERT: {a}")
            send_discord_alert(alerts, run_time)

        # Data quality flags (scrape validation — fires separately)
        flags = check_data_quality(results, prev)
        if flags:
            for f in flags:
                log.warning(f"DATA FLAG: {f}")
            send_discord_flags(flags, run_time)


if __name__ == '__main__':
    main()
